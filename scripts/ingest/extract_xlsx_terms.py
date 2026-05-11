"""Extract the BIOL 1592 key-terms XLSX into per-unit JSON term lists.

Input:  materials/1592 key terms for Units 1-17.xlsx
        Single sheet, 17 columns (one per unit), row 1 is the header ("Unit 1" ...
        through "Unit 17"), rows 2+ are term values. Empty cells skipped.

Output: content/terms.json
        {
          "course": "BIOL 1592",
          "units": {
            "1": ["anion", "atom", "atomic number", ...],
            ...
          }
        }

Term normalization is intentionally minimal here. We preserve casing and
whitespace as the prof wrote it (e.g. "ADP" vs "atomic number"). Synonym /
plural / hyphen normalization for FITB grading happens at quiz time, not here.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "materials" / "1592 key terms for Units 1-17.xlsx"
OUT = ROOT / "content" / "terms.json"

HEADER_RE = re.compile(r"^\s*Unit\s+(\d+)\s*$", re.IGNORECASE)


def main() -> int:
    if not SRC.exists():
        print(f"Not found: {SRC}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active

    # Map column index -> unit number from row 1.
    unit_for_col: dict[int, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        m = HEADER_RE.match(str(cell.value))
        if m:
            unit_for_col[col_idx] = int(m.group(1))

    if not unit_for_col:
        print("No 'Unit N' headers found in row 1.", file=sys.stderr)
        return 1

    units: dict[str, list[str]] = {str(u): [] for u in unit_for_col.values()}
    seen_per_unit: dict[str, set[str]] = {str(u): set() for u in unit_for_col.values()}

    for row in ws.iter_rows(min_row=2, values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            unit = unit_for_col.get(cell.column)
            if unit is None:
                continue
            term = str(cell.value).strip()
            if not term:
                continue
            key = term.lower()
            if key in seen_per_unit[str(unit)]:
                continue
            seen_per_unit[str(unit)].add(key)
            units[str(unit)].append(term)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as fp:
        json.dump(
            {"course": "BIOL 1592", "units": units},
            fp,
            ensure_ascii=False,
            indent=2,
        )

    total = sum(len(v) for v in units.values())
    for unit_no in sorted(units.keys(), key=int):
        print(f"  Unit {int(unit_no):>2}: {len(units[unit_no]):>3} terms")
    print(f"\n{total} terms across {len(units)} units → {OUT.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
